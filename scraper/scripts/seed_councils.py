#!/usr/bin/env python3
"""
One-off script to seed all ~537 Australian councils from state government directories.

Usage:
    uv run python seed_councils.py                    # print SQL to stdout
    uv run python seed_councils.py --output migration  # write V3__seed_councils.sql
    uv run python seed_councils.py --output db         # insert directly into the database

Sources:
    NSW: https://www.olg.nsw.gov.au/public/local-government-directory/
    VIC: https://www.vic.gov.au/local-government-contacts-and-information (Excel download)
    QLD: https://www.dlgwv.qld.gov.au/ (dropdown + JSON API)
    WA:  https://mycouncil.wa.gov.au/Home/GetGeoData (JSON API) + profile pages
    SA:  https://www.lga.sa.gov.au/sa-councils/councils-listing (Playwright)
    TAS: https://en.wikipedia.org/wiki/Local_government_areas_of_Tasmania (requests)
    NT:  https://nt.gov.au/community/.../find-your-council (requests)
    ACT: hard-coded (1 council)
"""

import argparse
import re
import sys
import tempfile
from pathlib import Path
from urllib.parse import urlparse

import psycopg2
import requests
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright

sys.path.insert(0, str(Path(__file__).parent.parent))
from config import settings

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _load_overrides() -> dict:
    path = Path(__file__).parent.parent / "council_overrides.yaml"
    if not path.exists():
        return {}
    import yaml

    return yaml.safe_load(path.read_text()) or {}


def slugify(name: str) -> str:
    s = name.lower().strip()
    s = re.sub(r"[''`]", "", s)  # apostrophes
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[\s_]+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def _esc(value: str | None) -> str:
    """Escape a string for SQL single-quoted literal."""
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


_WIKI_BASE = "https://en.wikipedia.org"
_WIKI_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; au-recycling-seed/1.0; "
        "+https://australiarecycling.com.au)"
    )
}


_WIKI_SKIP_DOMAINS = {"web.archive.org", "lga.sa.gov.au", "id.com.au", "wikipedia.org"}


def _canonicalize_url(url: str | None) -> str | None:
    """Upgrade http:// to https:// and strip redundant default ports."""
    if not url:
        return None
    url = url.strip()
    if url.startswith("http://"):
        url = "https://" + url[7:]
    # Strip redundant :443 (e.g. https://host:443/ → https://host/)
    parsed = urlparse(url)
    if parsed.scheme == "https" and parsed.port == 443:
        url = parsed._replace(netloc=parsed.hostname).geturl()
    return url.rstrip("/")


def _wiki_official_website(wiki_path: str) -> str | None:
    """
    Fetch a Wikipedia article and return the council's website URL from the
    External links section.

    Australian council Wikipedia articles use varied link text:
    "Official website", "Council website", "Adelaide Hills Council website", etc.
    Strategy: take the first external link containing "website" in the text,
    skipping archive/directory links.
    """
    import time
    from urllib.parse import urlparse

    try:
        resp = requests.get(_WIKI_BASE + wiki_path, timeout=15, headers=_WIKI_HEADERS)
        if not resp.ok:
            return None
        soup = BeautifulSoup(resp.text, "lxml")
        for h in soup.select("h2, h3"):
            if "external" in h.get_text(strip=True).lower():
                ul = h.find_next("ul")
                if ul:
                    first_valid: str | None = None
                    for li in ul.select("li"):
                        a = li.find("a", href=True)
                        if not a:
                            continue
                        href = a["href"]
                        domain = urlparse(href).netloc.removeprefix("www.")
                        if any(skip in domain for skip in _WIKI_SKIP_DOMAINS):
                            continue
                        if not href.startswith("http"):
                            continue
                        # Skip non-HTML resources
                        path_lower = urlparse(href).path.lower()
                        if path_lower.endswith(".pdf") or "contentfile" in path_lower:
                            continue
                        text = a.get_text(strip=True).lower()
                        # Prefer explicit "website"/"official" label
                        if "website" in text or "official" in text:
                            return href.rstrip("/")
                        # Fall back to first valid external link
                        if first_valid is None:
                            first_valid = href.rstrip("/")
                    if first_valid:
                        return first_valid
                break
    except Exception:
        pass
    finally:
        time.sleep(0.2)
    return None


def to_sql_row(
    name: str, state: str, website: str | None, slug: str | None = None
) -> str:
    slug = slug or slugify(name)
    return f"  ({_esc(name)}, {_esc(slug)}, {_esc(state)}, {_esc(website)})"


# ---------------------------------------------------------------------------
# NSW — OLG Local Government Directory (server-rendered HTML)
# ---------------------------------------------------------------------------


def scrape_nsw() -> list[dict]:
    print("  Fetching NSW...", file=sys.stderr)
    url = settings.seed_nsw_url
    resp = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")

    # The page has a two-level accordion. Top level has categories:
    # "Council" (128), "County Council", "Joint Organisation", etc.
    # We only want the "Council" category.
    council_section = None
    for item in soup.select(".accordion > .accordion-item"):
        title = item.select_one(".accordion-title")
        if title and title.get_text(strip=True) == "Council":
            council_section = item
            break

    if not council_section:
        print("    Could not find 'Council' section", file=sys.stderr)
        return []

    councils = []
    for item in council_section.select(".accordion-child .accordion-item"):
        name_el = item.select_one(".accordion-title")
        if not name_el:
            continue
        name = name_el.get_text(strip=True)

        website = None
        for b in item.select("b"):
            if "web" in b.get_text(strip=True).lower():
                a = b.find_next("a", href=True)
                if a:
                    href = a["href"].strip()
                    if not href.startswith("http"):
                        href = "https://" + href
                    website = href.rstrip("/")
                break

        councils.append({"name": name, "state": "NSW", "website": website})

    print(f"    NSW: {len(councils)} councils", file=sys.stderr)
    return councils


# ---------------------------------------------------------------------------
# VIC — vic.gov.au Excel download
# ---------------------------------------------------------------------------


def scrape_vic(page) -> list[dict]:
    print("  Fetching VIC...", file=sys.stderr)
    try:
        import openpyxl
    except ImportError:
        print(
            "    openpyxl not available — install with: uv add openpyxl",
            file=sys.stderr,
        )
        return []

    # Find the Excel download link via Playwright (page is JS-rendered)
    page.goto(
        settings.seed_vic_url,
        wait_until="domcontentloaded",
        timeout=30000,
    )
    xlsx_url = None
    for link in page.query_selector_all("a[href]"):
        href = link.get_attribute("href") or ""
        if ".xlsx" in href.lower():
            xlsx_url = (
                href if href.startswith("http") else "https://www.vic.gov.au" + href
            )
            break

    if not xlsx_url:
        print("    Could not find VIC Excel URL", file=sys.stderr)
        return []

    print(f"    Downloading VIC Excel: {xlsx_url}", file=sys.stderr)
    resp = requests.get(xlsx_url, timeout=60, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(resp.content)
        tmp_path = f.name

    wb = openpyxl.load_workbook(tmp_path, data_only=True)
    # The directory sheet has councils in repeating 5-row blocks:
    #   Row 1: [None, 'Council Name', ...]
    #   Row 2: address
    #   Row 3: suburb / postcode
    #   Row 4: [None, None, None, None, 'Email:', None, 'email@...']
    #   Row 5: [None, None, None, None, 'www:', None, 'www.council.vic.gov.au']
    ws = wb["DIRECTORY"] if "DIRECTORY" in wb.sheetnames else wb.active

    # Collect names (col B non-None, col A/C/D None) and websites (col E == 'www:')
    name_rows: list[tuple[int, str]] = []  # (row_num, name)
    web_rows: list[tuple[int, str]] = []  # (row_num, url)

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        col_b = str(row[1]).strip() if row[1] else None
        col_e = str(row[4]).strip().lower() if row[4] else ""
        col_g = str(row[6]).strip() if row[6] else None

        # Name row: col B has text, cols A, C, D are empty, no comma/digit (not an address)
        if col_b and not row[0] and not (row[2] or row[3]):
            if (
                _COUNCIL_KEYWORDS.search(col_b)
                and "," not in col_b
                and not re.search(r"\d", col_b)
            ):
                name_rows.append((i, col_b))

        # Website row: col E is 'www:' and col G has the URL
        if col_e == "www:" and col_g:
            url = col_g if col_g.startswith("http") else "https://" + col_g
            web_rows.append((i, url))

    # Match each name to the closest following website row
    councils = []
    web_idx = 0
    for name_row_num, name in name_rows:
        while web_idx < len(web_rows) and web_rows[web_idx][0] < name_row_num:
            web_idx += 1
        website = web_rows[web_idx][1] if web_idx < len(web_rows) else None
        councils.append({"name": name, "state": "VIC", "website": website})

    print(f"    VIC: {len(councils)} councils", file=sys.stderr)
    return councils


# ---------------------------------------------------------------------------
# QLD — qld.gov.au council directory
# ---------------------------------------------------------------------------


def scrape_qld(_page=None) -> list[dict]:
    """
    QLD directory page embeds council IDs in a dropdown, then loads details
    per council via a JSON API endpoint.
    """
    print("  Fetching QLD...", file=sys.stderr)
    import time

    directory_url = settings.seed_qld_url
    api_url = settings.seed_qld_api_url
    headers = {"User-Agent": "Mozilla/5.0"}

    resp = requests.get(directory_url, timeout=30, headers=headers)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")

    options = [
        (o["value"], o.get_text(strip=True))
        for o in soup.select("#councilSelectDropdown option")
        if o.get("value")
    ]

    # "Weipa Town" is a town authority, not an LGA — exclude it
    _EXCLUDE = {"Weipa Town"}

    councils = []
    for code, fallback_name in options:
        r = requests.get(api_url + code, timeout=10, headers=headers)
        data = r.json().get("matrix", {})
        name = data.get("name") or fallback_name
        if name in _EXCLUDE:
            continue
        website = (data.get("contact_url") or "").rstrip("/") or None
        councils.append({"name": name, "state": "QLD", "website": website})
        time.sleep(0.1)

    print(f"    QLD: {len(councils)} councils", file=sys.stderr)
    return councils


# ---------------------------------------------------------------------------
# WA — mycouncil.wa.gov.au (JSON API + profile pages)
# ---------------------------------------------------------------------------


def scrape_wa(_page=None) -> list[dict]:
    """
    Step 1: GET /Home/GetGeoData — returns JSON with LGName and LGID.
    Step 2: GET /Council/ViewCouncil/{LGID} — extract council website link.
    """
    import time

    print("  Fetching WA...", file=sys.stderr)
    base = settings.seed_wa_url.rstrip("/")
    headers = {"User-Agent": "Mozilla/5.0"}

    resp = requests.get(f"{base}/Home/GetGeoData", timeout=30, headers=headers)
    resp.raise_for_status()
    data = resp.json()

    # Christmas Island and Cocos (Keeling) Islands are federal territories,
    # not WA local governments. Narrogin appears twice in the API — deduped here.
    _WA_EXCLUDE = {"Christmas Island", "Cocos (Keeling) Islands"}
    seen_wa: set[str] = set()

    councils = []
    for item in data:
        name = (item.get("LGName") or "").strip()
        lg_id = item.get("LGID")
        if not name or not lg_id or name in _WA_EXCLUDE or name in seen_wa:
            continue
        seen_wa.add(name)

        # Fetch profile page to get the council's own website via the
        # explicit "Website" field (icon-earth bullet row).
        website = None
        try:
            profile = requests.get(
                f"{base}/Council/ViewCouncil/{lg_id}", timeout=15, headers=headers
            )
            if profile.ok:
                psoup = BeautifulSoup(profile.text, "lxml")
                icon = psoup.select_one(".bullet-icon span[title='Website']")
                if icon:
                    row = icon.find_parent("div", class_="col-xs-12")
                    a = row and row.select_one(".bullet-text a[href]")
                    if a:
                        website = a["href"].rstrip("/")
        except Exception:
            pass

        councils.append({"name": name, "state": "WA", "website": website})
        time.sleep(0.15)

    print(f"    WA: {len(councils)} councils", file=sys.stderr)
    return councils


# ---------------------------------------------------------------------------
# SA — Wikipedia (lga.sa.gov.au is behind Cloudflare)
# ---------------------------------------------------------------------------


def scrape_sa(_page=None) -> list[dict]:
    """
    lga.sa.gov.au is behind Cloudflare and blocks all automated access.
    Wikipedia's LGA list is accessible via requests and covers all 68 active SA councils
    across 7 wikitables (one per region).
    """
    print("  Fetching SA...", file=sys.stderr)
    url = settings.seed_sa_url
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (compatible; au-recycling-seed/1.0; "
            "+https://australiarecycling.com.au)"
        )
    }
    resp = requests.get(url, timeout=30, headers=headers)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")

    tables = soup.select("table.wikitable")
    if not tables:
        print("    SA: no wikitables found", file=sys.stderr)
        return []

    # Aboriginal land councils — not standard LGAs, excluded from the 68
    _SA_EXCLUDE = {
        "Aṉangu Pitjantjatjara Yankunytjatjara",
        "Maralinga Tjarutja",
        "Roxby Council",
    }

    seen: set[str] = set()
    councils = []

    for table in tables:
        for row in table.select("tr"):
            cells = row.select("td")
            if not cells:
                continue
            name = cells[0].get_text(strip=True)
            if not name or len(name) < 3 or name in seen or name in _SA_EXCLUDE:
                continue
            seen.add(name)

            wiki_link = cells[0].find("a", href=True)
            website = _wiki_official_website(wiki_link["href"]) if wiki_link else None
            councils.append({"name": name, "state": "SA", "website": website})

    print(f"    SA: {len(councils)} councils", file=sys.stderr)
    return councils


# ---------------------------------------------------------------------------
# TAS — Wikipedia (accessible via requests; 29 councils in wikitable)
# ---------------------------------------------------------------------------


def scrape_tas(_page=None) -> list[dict]:
    """
    Wikipedia's Local_government_areas_of_Tasmania is accessible via requests.
    The wikitable lists LGA short names (e.g. "Break O'Day", "Brighton").
    We trust all non-header rows in the single wikitable — no keyword filter needed.
    """
    print("  Fetching TAS...", file=sys.stderr)
    url = settings.seed_tas_url
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (compatible; au-recycling-seed/1.0; "
            "+https://australiarecycling.com.au)"
        )
    }
    resp = requests.get(url, timeout=30, headers=headers)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")

    tables = soup.select("table.wikitable")
    if not tables:
        print("    TAS: no wikitable found", file=sys.stderr)
        return []

    seen: set[str] = set()
    councils = []

    for row in tables[0].select("tr"):
        cells = row.select("td")
        if not cells:
            continue
        name = cells[0].get_text(strip=True)
        if not name or len(name) < 3 or name in seen:
            continue
        seen.add(name)

        wiki_link = cells[0].find("a", href=True)
        website = _wiki_official_website(wiki_link["href"]) if wiki_link else None
        councils.append({"name": name, "state": "TAS", "website": website})

    print(f"    TAS: {len(councils)} councils", file=sys.stderr)
    return councils


# ---------------------------------------------------------------------------
# NT — nt.gov.au (Playwright — regional councils are in JS accordion)
# ---------------------------------------------------------------------------

# NT has 18 councils (5 city/town, 3 shire/community government, 10 regional).
# The server-rendered HTML only shows 8 (municipal + shire).
# Regional councils are in a JavaScript accordion that requires Playwright.

# Tennant Creek is the seat of Barkly Regional Council, not a separate LGA.
_NT_EXCLUDE = {"Tennant Creek Town Council"}

_NT_FALLBACK = [
    ("Alice Springs Town Council", "https://www.alicesprings.nt.gov.au"),
    ("Barkly Regional Council", "https://www.barkly.nt.gov.au"),
    ("Belyuen Community Government Council", "https://www.belyuen.nt.gov.au"),
    ("Central Desert Regional Council", "https://www.centraldesert.nt.gov.au"),
    ("City of Darwin", "https://www.darwin.nt.gov.au"),
    ("City of Palmerston", "https://www.palmerston.nt.gov.au"),
    ("Coomalie Community Government Council", "https://www.coomalie.nt.gov.au"),
    ("East Arnhem Regional Council", "https://www.eastarnhem.nt.gov.au"),
    ("Groote Archipelago Regional Council", None),
    ("Katherine Town Council", "https://www.katherine.nt.gov.au"),
    ("Litchfield Council", "https://www.litchfield.nt.gov.au"),
    ("MacDonnell Regional Council", "https://www.macdonnell.nt.gov.au"),
    ("Roper Gulf Regional Council", "https://www.ropergulf.nt.gov.au"),
    ("Tiwi Islands Regional Council", "https://www.tiwiislands.nt.gov.au"),
    ("Victoria Daly Regional Council", "https://www.victoriadaly.nt.gov.au"),
    ("Wagait Shire Council", "https://www.wagait.nt.gov.au"),
    ("West Arnhem Regional Council", "https://www.westarnhem.nt.gov.au"),
    ("West Daly Regional Council", None),
]


def scrape_nt(page=None) -> list[dict]:
    """
    Use Playwright to fully render the nt.gov.au page (regional councils are
    in a JS accordion). Falls back to a hard-coded list if Playwright fails.
    """
    print("  Fetching NT...", file=sys.stderr)
    url = settings.seed_nt_url

    if page is not None:
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            # Expand all accordion sections
            page.wait_for_timeout(2000)
            for btn in page.query_selector_all(
                "button[aria-expanded='false'], .accordion-toggle"
            ):
                try:
                    btn.click()
                except Exception:
                    pass
            page.wait_for_timeout(1000)
            html = page.content()
            soup = BeautifulSoup(html, "lxml")

            seen: set[str] = set()
            councils = []
            for a in soup.select("a[href]"):
                name = re.sub(r"\s+website$", "", a.get_text(strip=True), flags=re.I)
                href = a.get("href", "")
                if not name or len(name) < 4 or name in seen or name in _NT_EXCLUDE:
                    continue
                if not _COUNCIL_KEYWORDS.search(name):
                    continue
                if not href.startswith("http") or "nt.gov.au/community" in href:
                    continue
                seen.add(name)
                councils.append(
                    {"name": name, "state": "NT", "website": href.rstrip("/")}
                )

            if len(councils) >= 10:
                print(f"    NT: {len(councils)} councils (Playwright)", file=sys.stderr)
                return councils
        except Exception as e:
            print(f"    NT Playwright failed: {e}", file=sys.stderr)

    print("    NT: using hard-coded list (18 councils)", file=sys.stderr)
    result = [
        {"name": n, "state": "NT", "website": w}
        for n, w in _NT_FALLBACK
        if n not in _NT_EXCLUDE
    ]
    print(f"    NT: {len(result)} councils", file=sys.stderr)
    return result


# ---------------------------------------------------------------------------
# ACT — single council
# ---------------------------------------------------------------------------


def scrape_act() -> list[dict]:
    print("  Using hard-coded ACT...", file=sys.stderr)
    return [
        {"name": "ACT Government", "state": "ACT", "website": "https://www.act.gov.au"}
    ]


_COUNCIL_KEYWORDS = re.compile(
    r"\b(council|shire|city|town|municipality|municipal|regional|district|borough|"
    r"aboriginal|community government)\b",
    re.I,
)

# Canonical name overrides — source directories sometimes list councils under
# inverted or abbreviated names. Map scraped name → canonical name so the slug
# stays stable across re-seeds and matches what is already in the DB.
_NAME_OVERRIDES: dict[str, str] = {
    # NSW OLG lists as "Council of the City of Sydney" / inverted form
    "Sydney, Council of the City of": "City of Sydney",
    "Council of the City of Sydney": "City of Sydney",
    # VIC Excel uses "Melbourne City Council"; official brand is "City of Melbourne"
    "Melbourne City Council": "City of Melbourne",
    # VIC Excel uses "Port Phillip City Council"; existing slug is port-phillip-council
    "Port Phillip City Council": "Port Phillip Council",
    # WA API returns bare suburb name "Perth"; official name is "City of Perth"
    "Perth": "City of Perth",
}


# ---------------------------------------------------------------------------
# SQL generation
# ---------------------------------------------------------------------------


def _prepare_councils(all_councils: list[dict]) -> list[dict]:
    """Apply name overrides, dedup by slug, and canonicalize URLs."""
    seen_slugs: set[str] = set()
    unique = []
    for c in all_councils:
        name = _NAME_OVERRIDES.get(c["name"], c["name"])
        slug = c.get("slug") or slugify(name)
        if slug and slug not in seen_slugs:
            seen_slugs.add(slug)
            unique.append({**c, "name": name, "slug": slug})

    print(f"  Canonicalizing {len(unique)} website URLs...", file=sys.stderr)
    for c in unique:
        c["website"] = _canonicalize_url(c["website"])

    return unique


def generate_sql(all_councils: list[dict]) -> tuple[str, list[dict]]:
    """Return (sql, prepared_councils). prepared_councils is deduped and canonicalized."""
    unique = _prepare_councils(all_councils)
    rows = [to_sql_row(c["name"], c["state"], c["website"], c["slug"]) for c in unique]
    rows_sql = ",\n".join(rows)

    sql = f"""\
-- Auto-generated by scraper/seed_councils.py
-- {len(unique)} councils across all Australian states and territories
INSERT INTO councils (name, slug, state, website) VALUES
{rows_sql}
ON CONFLICT (slug) DO UPDATE SET
  website = COALESCE(EXCLUDED.website, councils.website);
"""
    return sql, unique


# ---------------------------------------------------------------------------
# Output modes
# ---------------------------------------------------------------------------


def write_migration(sql: str) -> None:
    path = (
        Path(__file__).parent.parent.parent
        / "backend/src/main/resources/db/migration/V3__seed_councils.sql"
    )
    path.write_text(sql)
    print(f"Written: {path}", file=sys.stderr)


def _db_conn():
    import os
    from dotenv import load_dotenv

    load_dotenv(Path(__file__).parent.parent.parent / ".env")
    db_url = os.environ.get("DATABASE_URL", "")

    m = re.match(r"jdbc:postgresql://([^:/]+)(?::(\d+))?/(\w+)", db_url)
    if m:
        host, port, dbname = m.group(1), m.group(2) or "5432", m.group(3)
    else:
        host, port, dbname = "localhost", "5432", "recycling"

    return psycopg2.connect(
        host=host,
        port=port,
        dbname=dbname,
        user=os.environ.get("DATABASE_USERNAME", "recycling"),
        password=os.environ.get("DATABASE_PASSWORD", "recycling_dev"),
    )


def load_from_yaml(path: Path) -> list[dict]:
    """Load councils from a councils YAML file, skipping scraping entirely."""
    import yaml

    data = yaml.safe_load(path.read_text()) or []
    result = [
        {
            "name": c["name"],
            "slug": c["slug"],
            "state": c["state"],
            "website": c.get("website"),
        }
        for c in data
    ]
    print(f"Loaded {len(result)} councils from {path}", file=sys.stderr)
    return result


def write_to_db(sql: str, reset: bool = False) -> None:
    conn = _db_conn()
    with conn:
        with conn.cursor() as cur:
            if reset:
                cur.execute("TRUNCATE councils CASCADE")
                print(
                    "  Reset: truncated councils (cascades to suburbs, council_materials)",
                    file=sys.stderr,
                )
            cur.execute(sql)
    conn.close()
    print("Inserted into database.", file=sys.stderr)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description="Seed all Australian councils")
    parser.add_argument(
        "--output",
        choices=["stdout", "migration", "db"],
        default="stdout",
        help="stdout: print SQL | migration: write V3__seed_councils.sql | db: insert directly",
    )
    parser.add_argument(
        "--states",
        nargs="+",
        default=["NSW", "VIC", "QLD", "WA", "SA", "TAS", "NT", "ACT"],
        help="Limit to specific states (default: all)",
    )
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Truncate councils (cascades to suburbs, council_materials) before seeding. Only valid with --output db.",
    )
    parser.add_argument(
        "--from-file",
        metavar="PATH",
        help="Skip scraping; seed from the given councils YAML file instead.",
    )
    args = parser.parse_args()

    if args.reset and args.output != "db":
        parser.error("--reset is only valid with --output db")

    if args.from_file:
        all_councils = load_from_yaml(Path(args.from_file))
    else:
        all_councils = []

        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()

            state_scrapers = {
                "NSW": scrape_nsw,
                "VIC": lambda: scrape_vic(page),
                "QLD": scrape_qld,
                "WA": scrape_wa,
                "SA": scrape_sa,
                "TAS": scrape_tas,
                "NT": lambda: scrape_nt(page),
                "ACT": scrape_act,
            }

            for state in args.states:
                try:
                    councils = state_scrapers[state]()
                    all_councils.extend(councils)
                except Exception as e:
                    print(f"  ERROR scraping {state}: {e}", file=sys.stderr)

            browser.close()

    overrides = _load_overrides()
    for c in all_councils:
        slug = c.get("slug") or slugify(_NAME_OVERRIDES.get(c["name"], c["name"]))
        if slug in overrides and "website" in overrides[slug]:
            c["website"] = overrides[slug]["website"]

    sql, prepared = generate_sql(all_councils)

    print(f"\nTotal unique councils: {len(prepared)}", file=sys.stderr)

    if args.output == "stdout":
        print(sql)
    elif args.output == "migration":
        write_migration(sql)
    elif args.output == "db":
        write_to_db(sql, reset=args.reset)


if __name__ == "__main__":
    main()
