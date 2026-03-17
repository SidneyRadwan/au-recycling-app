#!/usr/bin/env python3
"""
One-off script to seed all ~537 Australian councils from state government directories.

Usage:
    uv run python seed_councils.py                    # print SQL to stdout
    uv run python seed_councils.py --output migration  # write V3__seed_councils.sql
    uv run python seed_councils.py --output db         # insert directly into the database

Sources:
    NSW: https://www.olg.nsw.gov.au/public/local-government-directory/
    VIC: https://www.vic.gov.au/local-government-contacts-and-information (Excel download)
    QLD: https://www.dlgwv.qld.gov.au/ (dropdown + JSON API)
    WA:  https://mycouncil.wa.gov.au/Home/GetGeoData (JSON API) + profile pages
    SA:  https://www.lga.sa.gov.au/sa-councils/councils-listing (Playwright)
    TAS: https://en.wikipedia.org/wiki/Local_government_areas_of_Tasmania (requests)
    NT:  https://nt.gov.au/community/.../find-your-council (requests)
    ACT: hard-coded (1 council)
"""

import argparse
import re
import sys
import tempfile
from pathlib import Path

import psycopg2
import requests
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright

from config import settings

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def slugify(name: str) -> str:
    s = name.lower().strip()
    s = re.sub(r"[''`]", "", s)  # apostrophes
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[\s_]+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def _esc(value: str | None) -> str:
    """Escape a string for SQL single-quoted literal."""
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


_WIKI_BASE = "https://en.wikipedia.org"
_WIKI_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; au-recycling-seed/1.0; "
        "+https://australiarecycling.com.au)"
    )
}


def _wiki_official_website(wiki_path: str) -> str | None:
    """
    Fetch a Wikipedia article and return the URL labelled 'Official website'
    in the External links section, or None if not found.
    """
    import time

    try:
        resp = requests.get(_WIKI_BASE + wiki_path, timeout=15, headers=_WIKI_HEADERS)
        if not resp.ok:
            return None
        soup = BeautifulSoup(resp.text, "lxml")
        for h in soup.select("h2, h3"):
            if "external" in h.get_text(strip=True).lower():
                ul = h.find_next("ul")
                if ul:
                    for li in ul.select("li"):
                        a = li.find("a", href=True)
                        if a and "official website" in a.get_text(strip=True).lower():
                            return a["href"].rstrip("/")
                break
    except Exception:
        pass
    finally:
        time.sleep(0.2)
    return None


def to_sql_row(name: str, state: str, website: str | None) -> str:
    slug = slugify(name)
    return f"  ({_esc(name)}, {_esc(slug)}, {_esc(state)}, {_esc(website)})"


# ---------------------------------------------------------------------------
# NSW — OLG Local Government Directory (server-rendered HTML)
# ---------------------------------------------------------------------------


def scrape_nsw() -> list[dict]:
    print("  Fetching NSW...", file=sys.stderr)
    url = settings.seed_nsw_url
    resp = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")

    # The page has a two-level accordion. Top level has categories:
    # "Council" (128), "County Council", "Joint Organisation", etc.
    # We only want the "Council" category.
    council_section = None
    for item in soup.select(".accordion > .accordion-item"):
        title = item.select_one(".accordion-title")
        if title and title.get_text(strip=True) == "Council":
            council_section = item
            break

    if not council_section:
        print("    Could not find 'Council' section", file=sys.stderr)
        return []

    councils = []
    for item in council_section.select(".accordion-child .accordion-item"):
        name_el = item.select_one(".accordion-title")
        if not name_el:
            continue
        name = name_el.get_text(strip=True)

        website = None
        for b in item.select("b"):
            if "web" in b.get_text(strip=True).lower():
                a = b.find_next("a", href=True)
                if a and a["href"].startswith("http"):
                    website = a["href"].rstrip("/")
                    break

        councils.append({"name": name, "state": "NSW", "website": website})

    print(f"    NSW: {len(councils)} councils", file=sys.stderr)
    return councils


# ---------------------------------------------------------------------------
# VIC — vic.gov.au Excel download
# ---------------------------------------------------------------------------


def scrape_vic(page) -> list[dict]:
    print("  Fetching VIC...", file=sys.stderr)
    try:
        import openpyxl
    except ImportError:
        print(
            "    openpyxl not available — install with: uv add openpyxl",
            file=sys.stderr,
        )
        return []

    # Find the Excel download link via Playwright (page is JS-rendered)
    page.goto(
        settings.seed_vic_url,
        wait_until="domcontentloaded",
        timeout=30000,
    )
    xlsx_url = None
    for link in page.query_selector_all("a[href]"):
        href = link.get_attribute("href") or ""
        if ".xlsx" in href.lower():
            xlsx_url = (
                href if href.startswith("http") else "https://www.vic.gov.au" + href
            )
            break

    if not xlsx_url:
        print("    Could not find VIC Excel URL", file=sys.stderr)
        return []

    print(f"    Downloading VIC Excel: {xlsx_url}", file=sys.stderr)
    resp = requests.get(xlsx_url, timeout=60, headers={"User-Agent": "Mozilla/5.0"})
    resp.raise_for_status()

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(resp.content)
        tmp_path = f.name

    wb = openpyxl.load_workbook(tmp_path, data_only=True)
    # The directory sheet has councils in repeating 5-row blocks:
    #   Row 1: [None, 'Council Name', ...]
    #   Row 2: address
    #   Row 3: suburb / postcode
    #   Row 4: [None, None, None, None, 'Email:', None, 'email@...']
    #   Row 5: [None, None, None, None, 'www:', None, 'www.council.vic.gov.au']
    ws = wb["DIRECTORY"] if "DIRECTORY" in wb.sheetnames else wb.active

    # Collect names (col B non-None, col A/C/D None) and websites (col E == 'www:')
    name_rows: list[tuple[int, str]] = []  # (row_num, name)
    web_rows: list[tuple[int, str]] = []  # (row_num, url)

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        col_b = str(row[1]).strip() if row[1] else None
        col_e = str(row[4]).strip().lower() if row[4] else ""
        col_g = str(row[6]).strip() if row[6] else None

        # Name row: col B has text, cols A, C, D are empty, no comma/digit (not an address)
        if col_b and not row[0] and not (row[2] or row[3]):
            if (
                _COUNCIL_KEYWORDS.search(col_b)
                and "," not in col_b
                and not re.search(r"\d", col_b)
            ):
                name_rows.append((i, col_b))

        # Website row: col E is 'www:' and col G has the URL
        if col_e == "www:" and col_g:
            url = col_g if col_g.startswith("http") else "https://" + col_g
            web_rows.append((i, url))

    # Match each name to the closest following website row
    councils = []
    web_idx = 0
    for name_row_num, name in name_rows:
        while web_idx < len(web_rows) and web_rows[web_idx][0] < name_row_num:
            web_idx += 1
        website = web_rows[web_idx][1] if web_idx < len(web_rows) else None
        councils.append({"name": name, "state": "VIC", "website": website})

    print(f"    VIC: {len(councils)} councils", file=sys.stderr)
    return councils


# ---------------------------------------------------------------------------
# QLD — qld.gov.au council directory
# ---------------------------------------------------------------------------


def scrape_qld(_page=None) -> list[dict]:
    """
    QLD directory page embeds council IDs in a dropdown, then loads details
    per council via a JSON API endpoint.
    """
    print("  Fetching QLD...", file=sys.stderr)
    import time

    directory_url = settings.seed_qld_url
    api_url = settings.seed_qld_api_url
    headers = {"User-Agent": "Mozilla/5.0"}

    resp = requests.get(directory_url, timeout=30, headers=headers)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")

    options = [
        (o["value"], o.get_text(strip=True))
        for o in soup.select("#councilSelectDropdown option")
        if o.get("value")
    ]

    # "Weipa Town" is a town authority, not an LGA — exclude it
    _EXCLUDE = {"Weipa Town"}

    councils = []
    for code, fallback_name in options:
        r = requests.get(api_url + code, timeout=10, headers=headers)
        data = r.json().get("matrix", {})
        name = data.get("name") or fallback_name
        if name in _EXCLUDE:
            continue
        website = (data.get("contact_url") or "").rstrip("/") or None
        councils.append({"name": name, "state": "QLD", "website": website})
        time.sleep(0.1)

    print(f"    QLD: {len(councils)} councils", file=sys.stderr)
    return councils


# ---------------------------------------------------------------------------
# WA — mycouncil.wa.gov.au (JSON API + profile pages)
# ---------------------------------------------------------------------------


def scrape_wa(_page=None) -> list[dict]:
    """
    Step 1: GET /Home/GetGeoData — returns JSON with LGName and LGID.
    Step 2: GET /Council/ViewCouncil/{LGID} — extract .wa.gov.au website link.
    """
    import time

    print("  Fetching WA...", file=sys.stderr)
    base = settings.seed_wa_url.rstrip("/")
    headers = {"User-Agent": "Mozilla/5.0"}

    resp = requests.get(f"{base}/Home/GetGeoData", timeout=30, headers=headers)
    resp.raise_for_status()
    data = resp.json()

    # Christmas Island and Cocos (Keeling) Islands are federal territories,
    # not WA local governments. Narrogin appears twice in the API — deduped here.
    _WA_EXCLUDE = {"Christmas Island", "Cocos (Keeling) Islands"}
    seen_wa: set[str] = set()

    councils = []
    for item in data:
        name = (item.get("LGName") or "").strip()
        lg_id = item.get("LGID")
        if not name or not lg_id or name in _WA_EXCLUDE or name in seen_wa:
            continue
        seen_wa.add(name)

        # Fetch profile page to get the council's own website
        website = None
        try:
            profile = requests.get(
                f"{base}/Council/ViewCouncil/{lg_id}", timeout=15, headers=headers
            )
            if profile.ok:
                psoup = BeautifulSoup(profile.text, "lxml")
                for a in psoup.select("a[href]"):
                    href = a["href"].rstrip("/")
                    if (
                        href.startswith("http")
                        and ".wa.gov.au" in href
                        and "mycouncil.wa.gov.au" not in href
                    ):
                        website = href
                        break
        except Exception:
            pass

        councils.append({"name": name, "state": "WA", "website": website})
        time.sleep(0.15)

    print(f"    WA: {len(councils)} councils", file=sys.stderr)
    return councils


# ---------------------------------------------------------------------------
# SA — Wikipedia (lga.sa.gov.au is behind Cloudflare)
# ---------------------------------------------------------------------------


def scrape_sa(_page=None) -> list[dict]:
    """
    lga.sa.gov.au is behind Cloudflare and blocks all automated access.
    Wikipedia's LGA list is accessible via requests and covers all 68 active SA councils
    across 7 wikitables (one per region).
    """
    print("  Fetching SA...", file=sys.stderr)
    url = settings.seed_sa_url
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (compatible; au-recycling-seed/1.0; "
            "+https://australiarecycling.com.au)"
        )
    }
    resp = requests.get(url, timeout=30, headers=headers)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")

    tables = soup.select("table.wikitable")
    if not tables:
        print("    SA: no wikitables found", file=sys.stderr)
        return []

    # Aboriginal land councils — not standard LGAs, excluded from the 68
    _SA_EXCLUDE = {
        "Aṉangu Pitjantjatjara Yankunytjatjara",
        "Maralinga Tjarutja",
        "Roxby Council",
    }

    seen: set[str] = set()
    councils = []

    for table in tables:
        for row in table.select("tr"):
            cells = row.select("td")
            if not cells:
                continue
            name = cells[0].get_text(strip=True)
            if not name or len(name) < 3 or name in seen or name in _SA_EXCLUDE:
                continue
            seen.add(name)

            wiki_link = cells[0].find("a", href=True)
            website = _wiki_official_website(wiki_link["href"]) if wiki_link else None
            councils.append({"name": name, "state": "SA", "website": website})

    print(f"    SA: {len(councils)} councils", file=sys.stderr)
    return councils


# ---------------------------------------------------------------------------
# TAS — Wikipedia (accessible via requests; 29 councils in wikitable)
# ---------------------------------------------------------------------------


def scrape_tas(_page=None) -> list[dict]:
    """
    Wikipedia's Local_government_areas_of_Tasmania is accessible via requests.
    The wikitable lists LGA short names (e.g. "Break O'Day", "Brighton").
    We trust all non-header rows in the single wikitable — no keyword filter needed.
    """
    print("  Fetching TAS...", file=sys.stderr)
    url = settings.seed_tas_url
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (compatible; au-recycling-seed/1.0; "
            "+https://australiarecycling.com.au)"
        )
    }
    resp = requests.get(url, timeout=30, headers=headers)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")

    tables = soup.select("table.wikitable")
    if not tables:
        print("    TAS: no wikitable found", file=sys.stderr)
        return []

    seen: set[str] = set()
    councils = []

    for row in tables[0].select("tr"):
        cells = row.select("td")
        if not cells:
            continue
        name = cells[0].get_text(strip=True)
        if not name or len(name) < 3 or name in seen:
            continue
        seen.add(name)

        wiki_link = cells[0].find("a", href=True)
        website = _wiki_official_website(wiki_link["href"]) if wiki_link else None
        councils.append({"name": name, "state": "TAS", "website": website})

    print(f"    TAS: {len(councils)} councils", file=sys.stderr)
    return councils


# ---------------------------------------------------------------------------
# NT — nt.gov.au (Playwright — regional councils are in JS accordion)
# ---------------------------------------------------------------------------

# NT has 5 municipal, 3 shire, and 10 regional councils = 18 total.
# The server-rendered HTML only shows 8 (municipal + shire).
# Regional councils are in a JavaScript accordion that requires Playwright.
_NT_FALLBACK = [
    ("Alice Springs Town Council", "https://www.alicesprings.nt.gov.au"),
    ("Barkly Regional Council", "https://www.barkly.nt.gov.au"),
    ("Belyuen Community Government Council", "https://www.belyuen.nt.gov.au"),
    ("Central Desert Regional Council", "https://www.centraldesert.nt.gov.au"),
    ("City of Darwin", "https://www.darwin.nt.gov.au"),
    ("City of Palmerston", "https://www.palmerston.nt.gov.au"),
    ("Coomalie Community Government Council", "https://www.coomalie.nt.gov.au"),
    ("East Arnhem Regional Council", "https://www.eastarnhem.nt.gov.au"),
    ("Groote Archipelago Regional Council", None),
    ("Katherine Town Council", "https://www.katherine.nt.gov.au"),
    ("Litchfield Council", "https://www.litchfield.nt.gov.au"),
    ("MacDonnell Regional Council", "https://www.macdonnell.nt.gov.au"),
    ("Roper Gulf Regional Council", "https://www.ropergulf.nt.gov.au"),
    ("Tennant Creek Town Council", "https://www.tennantcreek.nt.gov.au"),
    ("Tiwi Islands Regional Council", "https://www.tiwiislands.nt.gov.au"),
    ("Victoria Daly Regional Council", "https://www.victoriadaly.nt.gov.au"),
    ("Wagait Shire Council", "https://www.wagait.nt.gov.au"),
    ("West Arnhem Regional Council", "https://www.westarnhem.nt.gov.au"),
    ("West Daly Regional Council", None),
]


def scrape_nt(page=None) -> list[dict]:
    """
    Use Playwright to fully render the nt.gov.au page (regional councils are
    in a JS accordion). Falls back to a hard-coded list if Playwright fails.
    """
    print("  Fetching NT...", file=sys.stderr)
    url = settings.seed_nt_url

    if page is not None:
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            # Expand all accordion sections
            page.wait_for_timeout(2000)
            for btn in page.query_selector_all(
                "button[aria-expanded='false'], .accordion-toggle"
            ):
                try:
                    btn.click()
                except Exception:
                    pass
            page.wait_for_timeout(1000)
            html = page.content()
            soup = BeautifulSoup(html, "lxml")

            seen: set[str] = set()
            councils = []
            for a in soup.select("a[href]"):
                name = re.sub(r"\s+website$", "", a.get_text(strip=True), flags=re.I)
                href = a.get("href", "")
                if not name or len(name) < 4 or name in seen:
                    continue
                if not _COUNCIL_KEYWORDS.search(name):
                    continue
                if not href.startswith("http") or "nt.gov.au/community" in href:
                    continue
                seen.add(name)
                councils.append(
                    {"name": name, "state": "NT", "website": href.rstrip("/")}
                )

            if len(councils) >= 10:
                print(f"    NT: {len(councils)} councils (Playwright)", file=sys.stderr)
                return councils
        except Exception as e:
            print(f"    NT Playwright failed: {e}", file=sys.stderr)

    print("    NT: using hard-coded list (18 councils)", file=sys.stderr)
    result = [{"name": n, "state": "NT", "website": w} for n, w in _NT_FALLBACK]
    print(f"    NT: {len(result)} councils", file=sys.stderr)
    return result


# ---------------------------------------------------------------------------
# ACT — single council
# ---------------------------------------------------------------------------


def scrape_act() -> list[dict]:
    print("  Using hard-coded ACT...", file=sys.stderr)
    return [
        {"name": "ACT Government", "state": "ACT", "website": "https://www.act.gov.au"}
    ]


_COUNCIL_KEYWORDS = re.compile(
    r"\b(council|shire|city|town|municipality|municipal|regional|district|borough|"
    r"aboriginal|community government)\b",
    re.I,
)


# ---------------------------------------------------------------------------
# SQL generation
# ---------------------------------------------------------------------------


def generate_sql(all_councils: list[dict]) -> str:
    # Deduplicate by slug
    seen_slugs: set[str] = set()
    unique = []
    for c in all_councils:
        slug = slugify(c["name"])
        if slug and slug not in seen_slugs:
            seen_slugs.add(slug)
            unique.append(c)

    rows = [to_sql_row(c["name"], c["state"], c["website"]) for c in unique]
    rows_sql = ",\n".join(rows)

    return f"""\
-- Auto-generated by scraper/seed_councils.py
-- {len(unique)} councils across all Australian states and territories
INSERT INTO councils (name, slug, state, website) VALUES
{rows_sql}
ON CONFLICT (slug) DO NOTHING;
"""


# ---------------------------------------------------------------------------
# Output modes
# ---------------------------------------------------------------------------


def write_migration(sql: str) -> None:
    path = (
        Path(__file__).parent.parent
        / "backend/src/main/resources/db/migration/V3__seed_councils.sql"
    )
    path.write_text(sql)
    print(f"Written: {path}", file=sys.stderr)


def write_to_db(sql: str) -> None:
    import os
    from dotenv import load_dotenv

    load_dotenv(Path(__file__).parent.parent / ".env")
    db_url = os.environ.get("DATABASE_URL", "")

    # Convert jdbc URL to psycopg2 DSN
    # jdbc:postgresql://host:port/dbname -> host=... port=... dbname=...
    m = re.match(r"jdbc:postgresql://([^:/]+)(?::(\d+))?/(\w+)", db_url)
    if m:
        host, port, dbname = m.group(1), m.group(2) or "5432", m.group(3)
    else:
        host, port, dbname = "localhost", "5432", "recycling"

    user = os.environ.get("DATABASE_USERNAME", "recycling")
    password = os.environ.get("DATABASE_PASSWORD", "recycling_dev")

    conn = psycopg2.connect(
        host=host, port=port, dbname=dbname, user=user, password=password
    )
    with conn:
        with conn.cursor() as cur:
            cur.execute(sql)
    conn.close()
    print("Inserted into database.", file=sys.stderr)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description="Seed all Australian councils")
    parser.add_argument(
        "--output",
        choices=["stdout", "migration", "db"],
        default="stdout",
        help="stdout: print SQL | migration: write V3__seed_councils.sql | db: insert directly",
    )
    parser.add_argument(
        "--states",
        nargs="+",
        default=["NSW", "VIC", "QLD", "WA", "SA", "TAS", "NT", "ACT"],
        help="Limit to specific states (default: all)",
    )
    args = parser.parse_args()

    all_councils: list[dict] = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        page = browser.new_page()

        state_scrapers = {
            "NSW": scrape_nsw,
            "VIC": lambda: scrape_vic(page),
            "QLD": scrape_qld,
            "WA": scrape_wa,
            "SA": scrape_sa,
            "TAS": scrape_tas,
            "NT": lambda: scrape_nt(page),
            "ACT": scrape_act,
        }

        for state in args.states:
            try:
                councils = state_scrapers[state]()
                all_councils.extend(councils)
            except Exception as e:
                print(f"  ERROR scraping {state}: {e}", file=sys.stderr)

        browser.close()

    sql = generate_sql(all_councils)

    print(f"\nTotal unique councils: {sql.count(chr(10) + '  (')}", file=sys.stderr)

    if args.output == "stdout":
        print(sql)
    elif args.output == "migration":
        write_migration(sql)
    elif args.output == "db":
        write_to_db(sql)


if __name__ == "__main__":
    main()
